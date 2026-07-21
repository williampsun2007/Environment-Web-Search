'''
Batch-mode helpers: turning a user's spreadsheet into validated search
requests, and turning search results into a formatted Excel workbook.

validate_batch_df() checks an uploaded DataFrame has the required columns
(Location, Start Date, End Date, Pollutant) and well-formed rows, returning
the valid rows plus a list of per-row error messages.

The remaining helpers (_style_header, _autosize_columns, _wb_to_bytes) style
and serialize the two-sheet output workbook (Summary + Sources) used by the
Batch Search tab. This module has no dependency on search_pipeline.py or on
Streamlit's session state - it only knows about DataFrames and openpyxl
workbooks.
'''

import pandas as pd
from openpyxl.styles import Font, PatternFill
import io

# Helpers for the batch feature
def validate_batch_df(df):
    # Return (rows: list[dict], errors: list[str]) from a raw DataFrame.
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    required = {"location", "start date", "end date", "pollutant"}
    missing = required - set(df.columns)
    if missing:
        return [], [f"Missing required columns: {', '.join(sorted(missing))}"]

    rows, errors = [], []
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed header + offset)
        raw_vals = [str(row.get(c, "")).strip() for c in ("location", "start date", "end date", "pollutant")]
        if not any(v for v in raw_vals if v not in ("nan", "")):
            continue  # silently skip fully blank rows
        loc = str(row["location"]).strip() if pd.notna(row["location"]) else ""
        poll = str(row["pollutant"]).strip() if pd.notna(row["pollutant"]) else ""
        if loc in ("nan", ""):
            loc = ""
        if poll in ("nan", ""):
            poll = ""
        if not loc or not poll:
            errors.append(f"Row {row_num}: missing {'location' if not loc else 'pollutant'}")
            continue
        try:
            start = pd.to_datetime(row["start date"]).date()
        except Exception:
            errors.append(f"Row {row_num}: invalid start date '{row['start date']}'")
            continue
        try:
            end = pd.to_datetime(row["end date"]).date()
        except Exception:
            errors.append(f"Row {row_num}: invalid end date '{row['end date']}'")
            continue
        if start > end:
            errors.append(f"Row {row_num}: start date {start} is after end date {end}")
            continue
        rows.append({"location": loc, "start_date": start, "end_date": end, "pollutant": poll})
    return rows, errors

def _style_header(ws):
    hdr_font = Font(bold = True, color = "FFFFFF")
    hdr_fill = PatternFill("solid", fgColor = "4472C4")
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = hdr_font
            cell.fill = hdr_fill
    ws.freeze_panes = "A2"

def _autosize_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default = 0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

def _wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()